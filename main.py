import json
import os
import openpyxl
from dotenv import load_dotenv
from google import genai
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()
client = genai.Client(api_key=os.environ.get("GEMINI_API_KEY"))

SYSTEM_PROMPT = """
Eres un Analista de Calidad de Datos de la Clínica MediSalud.
Recibirás una lista de registros de pacientes y deberás validar cada uno según estas reglas:

1. El nombre es obligatorio (no puede estar vacío).
2. La edad debe estar entre 0 y 120 años.
3. El correo debe contener el carácter '@'.
4. El teléfono debe tener exactamente 9 dígitos.
5. La ciudad es obligatoria (no puede estar vacía).
6. La fecha de nacimiento no puede ser una fecha futura.

Responde ÚNICAMENTE con un arreglo JSON válido, sin texto adicional, con este formato por cada registro:
[
  {
    "id": <número de registro>,
    "estado": "Válido" o "Inválido",
    "errores": "<errores separados por punto y coma, o 'Ninguno'>",
    "recomendacion": "<acciones correctivas separadas por punto y coma, o 'Registro listo para carga.'>"
  }
]
""".strip()

REGISTROS = [
    {
        "id": 1,
        "nombre": "Juan Pérez",
        "edad": 35,
        "correo": "juan@gmail.com",
        "telefono": "987654321",
        "ciudad": "Santiago",
        "fecha_nacimiento": "1990-05-10",
    },
    {
        "id": 2,
        "nombre": "María Soto",
        "edad": -5,
        "correo": "mariagmail.com",
        "telefono": "12345",
        "ciudad": "Valparaíso",
        "fecha_nacimiento": "2030-01-01",
    },
    {
        "id": 3,
        "nombre": "",
        "edad": 28,
        "correo": "pedro@hotmail.com",
        "telefono": "912345678",
        "ciudad": "",
        "fecha_nacimiento": "1997-03-15",
    },
]


def validar_registros(registros: list) -> list:
    contenido = json.dumps(registros, ensure_ascii=False, indent=2)
    prompt = f"{SYSTEM_PROMPT}\n\nValida estos registros:\n{contenido}"

    response = client.models.generate_content(
        model="gemini-2.0-flash",
        contents=prompt,
    )

    texto = response.text.strip()
    # Eliminar posibles bloques de código markdown (```json ... ```)
    if texto.startswith("```"):
        texto = texto.split("```")[1]
        if texto.startswith("json"):
            texto = texto[4:]
    texto = texto.strip()

    datos = json.loads(texto)
    if isinstance(datos, list):
        return datos
    return next(iter(datos.values()))


def exportar_excel(registros: list, resultados: list, path: str = "resultado_validacion.xlsx"):
    resultados_por_id = {r["id"]: r for r in resultados}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validación de Pacientes"

    encabezados = [
        "ID", "Nombre", "Edad", "Correo", "Teléfono", "Ciudad",
        "Fecha Nacimiento", "Estado", "Errores Detectados", "Recomendación"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2F5496")

    for col, titulo in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    fill_valido = PatternFill(fill_type="solid", fgColor="C6EFCE")
    fill_invalido = PatternFill(fill_type="solid", fgColor="FFC7CE")

    for fila, registro in enumerate(registros, start=2):
        resultado = resultados_por_id.get(registro["id"], {})
        valores = [
            registro["id"],
            registro["nombre"] or "(vacío)",
            registro["edad"],
            registro["correo"],
            registro["telefono"],
            registro["ciudad"] or "(vacía)",
            registro["fecha_nacimiento"],
            resultado.get("estado", ""),
            resultado.get("errores", ""),
            resultado.get("recomendacion", ""),
        ]
        fill = fill_valido if resultado.get("estado") == "Válido" else fill_invalido

        for col, valor in enumerate(valores, start=1):
            cell = ws.cell(row=fila, column=col, value=valor)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    anchos = [5, 18, 8, 25, 12, 15, 18, 10, 45, 55]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    wb.save(path)
    print(f"Resultado exportado a: {path}")


if __name__ == "__main__":
    print("Enviando registros al agente...")
    resultados = validar_registros(REGISTROS)

    for r in resultados:
        print(f"Registro {r['id']}: {r['estado']}")

    exportar_excel(REGISTROS, resultados)
